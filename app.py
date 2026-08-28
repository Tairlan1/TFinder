# -*- coding: utf-8 -*-
"""
app.py — веб-интерфейс поверх scraper_engine.py.

Запуск:
    pip install -r requirements.txt
    python app.py
    -> открыть http://127.0.0.1:5000 в браузере (обычном, не в окне автоматизации)
"""

import os
import time

from flask import Flask, jsonify, request, render_template, send_file

import scraper_engine as engine

app = Flask(__name__)

# ----------------------------------------------------------------------------
# Кэш чтения xlsx-файлов (тендеры) - файлы могут быть довольно большими,
# а UI может опрашивать список часто; перечитываем с диска, только если
# файл реально изменился с прошлого раза (по mtime).
# ----------------------------------------------------------------------------
_tenders_cache = {"mtime_main": None, "mtime_astana": None, "rows": []}


def _read_xlsx_rows(path: str, city_label: str):
    if not os.path.exists(path):
        return []
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    header = engine.HEADERS
    for i, values in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not values or all(v is None for v in values):
            continue
        row = {header[j]: (values[j] if j < len(values) else "") for j in range(len(header))}
        row["Город"] = city_label
        row["_row_id"] = f"{city_label}:{i}"
        row["TenderId"] = engine._extract_tender_id_from_url(row.get("Ссылка на объявление", ""))
        rows.append(row)
    wb.close()
    return rows


def get_tenders_rows(force=False):
    mtime_main = os.path.getmtime(engine.OUTPUT_XLSX) if os.path.exists(engine.OUTPUT_XLSX) else None
    mtime_astana = os.path.getmtime(engine.OUTPUT_XLSX_ASTANA) if os.path.exists(engine.OUTPUT_XLSX_ASTANA) else None

    if (not force and mtime_main == _tenders_cache["mtime_main"]
            and mtime_astana == _tenders_cache["mtime_astana"]):
        return _tenders_cache["rows"]

    rows = _read_xlsx_rows(engine.OUTPUT_XLSX, "Остальные")
    rows += _read_xlsx_rows(engine.OUTPUT_XLSX_ASTANA, "Астана")

    _tenders_cache["mtime_main"] = mtime_main
    _tenders_cache["mtime_astana"] = mtime_astana
    _tenders_cache["rows"] = rows
    return rows


# ----------------------------------------------------------------------------
# Страницы
# ----------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


# ----------------------------------------------------------------------------
# API: ключевые слова (профили тематик)
# ----------------------------------------------------------------------------

@app.route("/api/keyword-profiles", methods=["GET"])
def api_keyword_profiles_get():
    return jsonify(engine.get_profiles_state())


@app.route("/api/keyword-profiles/save", methods=["POST"])
def api_keyword_profiles_save():
    data = request.get_json(force=True, silent=True) or {}
    name = str(data.get("name") or "").strip()
    keywords = data.get("keywords") or {}
    if not name:
        return jsonify({"ok": False, "error": "Не указано имя профиля"}), 400
    for key in ("strong", "medium", "weak", "negative", "context"):
        if key not in keywords or not isinstance(keywords[key], list):
            return jsonify({"ok": False, "error": f"Категория '{key}' должна быть списком строк"}), 400
    try:
        state = engine.save_profile(name, keywords)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    return jsonify({"ok": True, **state})


@app.route("/api/keyword-profiles/delete", methods=["POST"])
def api_keyword_profiles_delete():
    data = request.get_json(force=True, silent=True) or {}
    name = str(data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "Не указано имя профиля"}), 400
    state = engine.delete_profile(name)
    return jsonify({"ok": True, **state})


@app.route("/api/keyword-profiles/rename", methods=["POST"])
def api_keyword_profiles_rename():
    data = request.get_json(force=True, silent=True) or {}
    old_name = str(data.get("old_name") or "").strip()
    new_name = str(data.get("new_name") or "").strip()
    if not old_name or not new_name:
        return jsonify({"ok": False, "error": "Не указано имя профиля"}), 400
    try:
        state = engine.rename_profile(old_name, new_name)
    except ValueError as e:
        return jsonify({"ok": False, "error": str(e)}), 400
    return jsonify({"ok": True, **state})


@app.route("/api/keyword-profiles/active", methods=["POST"])
def api_keyword_profiles_active():
    data = request.get_json(force=True, silent=True) or {}
    names = data.get("active")
    if not isinstance(names, list):
        return jsonify({"ok": False, "error": "'active' должен быть списком имён профилей"}), 400
    state = engine.set_active_profiles(names)
    return jsonify({"ok": True, **state})


# ----------------------------------------------------------------------------
# API: ИИ-проверка релевантности (опционально)
# ----------------------------------------------------------------------------

@app.route("/api/ai-settings", methods=["GET"])
def api_ai_settings_get():
    settings = engine.get_ai_settings()
    return jsonify({
        **settings,
        "gemini_ready": engine.ai_provider_ready("gemini"),
        "anthropic_installed": engine.ANTHROPIC_AVAILABLE,
        "anthropic_ready": engine.ai_provider_ready("anthropic"),
    })


@app.route("/api/ai-settings", methods=["POST"])
def api_ai_settings_post():
    data = request.get_json(force=True, silent=True) or {}
    settings = engine.save_ai_settings(data)
    return jsonify({"ok": True, **settings})


# ----------------------------------------------------------------------------
# API: таблица найденных тендеров
# ----------------------------------------------------------------------------

@app.route("/api/tenders", methods=["GET"])
def api_tenders():
    rows = get_tenders_rows()

    search = (request.args.get("search") or "").strip().lower()
    priority = request.args.get("priority") or ""
    it_type = request.args.get("it_type") or ""
    status = request.args.get("status") or ""
    city = request.args.get("city") or ""
    date_from = request.args.get("date_from") or ""
    date_to = request.args.get("date_to") or ""
    min_amount = request.args.get("min_amount") or ""
    viewed_filter = request.args.get("viewed") or ""   # "", "yes", "no"
    sort_by = request.args.get("sort_by") or "Начало приёма заявок"
    sort_dir = request.args.get("sort_dir") or "desc"
    page = max(1, int(request.args.get("page", 1)))
    page_size = min(500, max(1, int(request.args.get("page_size", 50))))

    viewed_ids = engine.get_viewed_ids()
    favorite_ids = engine.get_favorite_ids()
    for r in rows:
        r["Просмотрено"] = bool(r.get("TenderId")) and r["TenderId"] in viewed_ids
        r["Избранное"] = bool(r.get("TenderId")) and r["TenderId"] in favorite_ids
        r["Просрочен"] = engine.is_expired(r.get("Окончание приёма заявок", ""))
        r["ДнейОсталось"] = engine.days_left(r.get("Окончание приёма заявок", ""))

    def matches(row):
        if viewed_filter == "yes" and not row.get("Просмотрено"):
            return False
        if viewed_filter == "no" and row.get("Просмотрено"):
            return False
        if priority and row.get("Приоритет") != priority:
            return False
        if it_type and row.get("Тип IT") != it_type:
            return False
        if status and row.get("Статус") != status:
            return False
        if city and row.get("Город") != city:
            return False
        if date_from and str(row.get("Начало приёма заявок") or "") < date_from:
            return False
        if date_to and str(row.get("Начало приёма заявок") or "") > date_to + "~":
            return False
        if min_amount:
            try:
                amt = float(str(row.get("Сумма, тг.") or "0").replace(" ", "").replace("\u00a0", "").replace(",", "."))
                if amt < float(min_amount):
                    return False
            except ValueError:
                pass
        if search:
            haystack = " ".join(str(row.get(k, "")) for k in (
                "Название лота/объявления", "Заказчик", "Организатор",
                "Совпавшие IT-ключевые слова", "№ объявления",
            )).lower()
            if search not in haystack:
                return False
        return True

    filtered = [r for r in rows if matches(r)]

    reverse = sort_dir == "desc"
    filtered.sort(key=lambda r: str(r.get(sort_by, "")), reverse=reverse)

    total = len(filtered)
    start = (page - 1) * page_size
    page_rows = filtered[start:start + page_size]

    clean_rows = [{k: v for k, v in r.items() if not k.startswith("_")} for r in page_rows]

    return jsonify({
        "total": total,
        "page": page,
        "page_size": page_size,
        "rows": clean_rows,
        "facets": {
            "priority": sorted({r.get("Приоритет", "") for r in rows if r.get("Приоритет")}),
            "it_type": sorted({r.get("Тип IT", "") for r in rows if r.get("Тип IT")}),
            "status": sorted({r.get("Статус", "") for r in rows if r.get("Статус")}),
        },
    })


@app.route("/api/tenders/refresh", methods=["POST"])
def api_tenders_refresh():
    rows = get_tenders_rows(force=True)
    return jsonify({"ok": True, "total": len(rows)})


@app.route("/api/tenders/viewed", methods=["POST"])
def api_tenders_viewed():
    data = request.get_json(force=True, silent=True) or {}
    tender_id = str(data.get("id") or "").strip()
    viewed = bool(data.get("viewed"))
    if not tender_id:
        return jsonify({"ok": False, "error": "Не передан id тендера"}), 400
    ids = engine.set_viewed(tender_id, viewed)
    return jsonify({"ok": True, "viewed": tender_id in ids})


@app.route("/api/tenders/delete", methods=["POST"])
def api_tenders_delete():
    data = request.get_json(force=True, silent=True) or {}
    tender_id = str(data.get("id") or "").strip()
    if not tender_id:
        return jsonify({"ok": False, "error": "Не передан id тендера"}), 400
    deleted = engine.delete_tender_row(tender_id)
    if deleted:
        get_tenders_rows(force=True)
    return jsonify({"ok": deleted})


# ----------------------------------------------------------------------------
# API: избранное
# ----------------------------------------------------------------------------

@app.route("/api/favorites", methods=["GET"])
def api_favorites_get():
    favorites = engine.get_favorites()
    search = (request.args.get("search") or "").strip().lower()

    rows = []
    for tender_id, entry in favorites.items():
        row = dict(entry)
        row["TenderId"] = tender_id
        row["Просрочен"] = engine.is_expired(entry.get("Окончание приёма заявок", ""))
        row["ДнейОсталось"] = engine.days_left(entry.get("Окончание приёма заявок", ""))
        rows.append(row)

    if search:
        rows = [
            r for r in rows
            if search in " ".join(str(r.get(k, "")) for k in (
                "Название лота/объявления", "Заказчик", "Организатор", "Заметка", "№ объявления",
            )).lower()
        ]

    rows.sort(key=lambda r: r.get("Добавлено", ""), reverse=True)
    return jsonify({"total": len(rows), "rows": rows})


@app.route("/api/favorites", methods=["POST"])
def api_favorites_add():
    data = request.get_json(force=True, silent=True) or {}
    tender_id = str(data.get("id") or data.get("TenderId") or "").strip()
    if not tender_id:
        return jsonify({"ok": False, "error": "Не передан id тендера"}), 400
    favorites = engine.add_favorite(tender_id, data)
    return jsonify({"ok": True, "total": len(favorites)})


@app.route("/api/favorites/remove", methods=["POST"])
def api_favorites_remove():
    data = request.get_json(force=True, silent=True) or {}
    tender_id = str(data.get("id") or "").strip()
    if not tender_id:
        return jsonify({"ok": False, "error": "Не передан id тендера"}), 400
    favorites = engine.remove_favorite(tender_id)
    return jsonify({"ok": True, "total": len(favorites)})


@app.route("/api/favorites/note", methods=["POST"])
def api_favorites_note():
    data = request.get_json(force=True, silent=True) or {}
    tender_id = str(data.get("id") or "").strip()
    note = str(data.get("note") or "")
    if not tender_id:
        return jsonify({"ok": False, "error": "Не передан id тендера"}), 400
    favorites = engine.set_favorite_note(tender_id, note)
    if tender_id not in favorites:
        return jsonify({"ok": False, "error": "Тендер не найден в избранном"}), 404
    return jsonify({"ok": True})


@app.route("/api/favorites/export", methods=["GET"])
def api_favorites_export():
    if not os.path.exists(engine.FAVORITES_XLSX):
        engine._regenerate_favorites_xlsx(engine.get_favorites())
    return send_file(
        engine.FAVORITES_XLSX,
        as_attachment=True,
        download_name="izbrannye_tendery.xlsx",
    )


# ----------------------------------------------------------------------------
# API: пограничные кандидаты (borderline)
# ----------------------------------------------------------------------------

def _read_borderline_rows():
    path = engine.BORDERLINE_LOG_FILE
    if not os.path.exists(path):
        return []
    rows = []
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            parts = line.split("\t")
            if len(parts) < 5:
                continue

            ts, score_part, number_anno, url, title = parts[:5]
            try:
                score = int(score_part.replace("score=", ""))
            except ValueError:
                score = 0

            # Всё после title - самоописанные key=value поля (amount=,
            # end_date=, keywords=), в любом порядке/количестве. Самый
            # первый (исторический) формат лога тоже совместим: keywords=
            # уже тогда шёл с этим префиксом.
            extra = {}
            for p in parts[5:]:
                if "=" in p:
                    k, _, v = p.partition("=")
                    extra[k] = v
                else:
                    extra.setdefault("keywords", p)

            tender_id = engine._extract_tender_id_from_url(url)
            rows.append({
                "ts": ts, "score": score, "number_anno": number_anno,
                "url": url, "title": title,
                "amount": extra.get("amount", ""),
                "start_date": extra.get("start_date", ""),
                "end_date": extra.get("end_date", ""),
                "keywords": extra.get("keywords", ""),
                "tender_id": tender_id,
            })
    return rows


@app.route("/api/borderline", methods=["GET"])
def api_borderline():
    rows = _read_borderline_rows()

    min_score = request.args.get("min_score")
    search = (request.args.get("search") or "").strip().lower()
    page = max(1, int(request.args.get("page", 1)))
    page_size = min(500, max(1, int(request.args.get("page_size", 50))))

    if min_score:
        try:
            rows = [r for r in rows if r["score"] >= int(min_score)]
        except ValueError:
            pass
    if search:
        rows = [r for r in rows if search in (r["title"] + " " + r["keywords"]).lower()]

    rows.sort(key=lambda r: r["score"], reverse=True)

    total = len(rows)
    start = (page - 1) * page_size
    page_rows = rows[start:start + page_size]

    favorite_ids = engine.get_favorite_ids()
    for r in page_rows:
        r["Избранное"] = bool(r.get("tender_id")) and r["tender_id"] in favorite_ids
        r["Просрочен"] = engine.is_expired(r.get("end_date", ""))
        r["ДнейОсталось"] = engine.days_left(r.get("end_date", ""))

    return jsonify({"total": total, "page": page, "page_size": page_size, "rows": page_rows})


@app.route("/api/borderline/delete", methods=["POST"])
def api_borderline_delete():
    data = request.get_json(force=True, silent=True) or {}
    tender_id = str(data.get("id") or "").strip()
    if not tender_id:
        return jsonify({"ok": False, "error": "Не передан id тендера"}), 400
    removed = engine.delete_borderline_by_tender_id(tender_id)
    return jsonify({"ok": removed > 0, "removed": removed})


# ----------------------------------------------------------------------------
# API: управление сбором (запуск/остановка/статус/лог)
# ----------------------------------------------------------------------------

@app.route("/api/run/state", methods=["GET"])
def api_run_state():
    return jsonify(engine.runner.snapshot())


@app.route("/api/run/set_url", methods=["POST"])
def api_run_set_url():
    data = request.get_json(force=True, silent=True) or {}
    url = str(data.get("url") or "").strip()
    ok, info = engine.runner.set_search_url(url)
    return jsonify({"ok": ok, "info": info, **engine.runner.snapshot()})


@app.route("/api/run/set_auto_restart", methods=["POST"])
def api_run_set_auto_restart():
    data = request.get_json(force=True, silent=True) or {}
    engine.runner.set_auto_restart(bool(data.get("enabled")))
    return jsonify({"ok": True, **engine.runner.snapshot()})


@app.route("/api/run/launch_browser", methods=["POST"])
def api_run_launch_browser():
    ok = engine.runner.launch_browser()
    return jsonify({"ok": ok, **engine.runner.snapshot()})


@app.route("/api/run/confirm_ready", methods=["POST"])
def api_run_confirm_ready():
    ok, info = engine.runner.confirm_ready()
    return jsonify({"ok": ok, "info": info, **engine.runner.snapshot()})


@app.route("/api/run/start", methods=["POST"])
def api_run_start():
    ok, info = engine.runner.start_run()
    return jsonify({"ok": ok, "info": info, **engine.runner.snapshot()})


@app.route("/api/run/stop", methods=["POST"])
def api_run_stop():
    ok, info = engine.runner.stop_run()
    return jsonify({"ok": ok, "info": info, **engine.runner.snapshot()})


@app.route("/api/run/logs", methods=["GET"])
def api_run_logs():
    since = int(request.args.get("since", 0))
    lines = engine.get_log_since(since)
    return jsonify({"lines": lines})


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False, threaded=True)
